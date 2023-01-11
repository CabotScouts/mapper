import csv
from pathlib import Path
import urllib.parse

from openpyxl import load_workbook
import requests
import matplotlib.pyplot as pl
import cartopy.crs as ccrs
import cartopy.io.img_tiles as cimgt
import numpy


class PostcodeMapper(object):

    title = None
    endpoint = "http://api.getthedata.com/postcode/"
    coords = []
    csvfields = ["postcode", "longitude", "latitude"]
    cache = dict()
    cachefile = None
    recache = False
    xbins = 40  # Number of bins for histogram - smaller = less resolution
    ybins = 40
    tiledepth = 16
    markers = []  # list of numpy arrays [lat, long, colour] for each point to mark
    figure = None  # matplotlib figure

    def __init__(self, cache=None, **kwargs):
        cachefile = cache if cache else "mapper/postcodes.csv"
        self.cachefile = Path(cachefile)
        self.loadCache()
        self.title = kwargs.get("title", None)

    def loadCache(self):
        # Retrieves cached coordinates, if they exist
        if self.cachefile.is_file():
            with open(self.cachefile, "r") as c:
                d = csv.DictReader(c, fieldnames=self.csvfields)
                next(d, None)  # Skip header row
                for row in d:
                    self.cache[row["postcode"]] = (
                        float(row["longitude"]),
                        float(row["latitude"]),
                    )

    def saveCache(self):
        if self.recache:
            with open(self.cachefile, "w") as c:
                d = csv.DictWriter(c, fieldnames=self.csvfields)
                d.writeheader()
                for k, v in self.cache.items():
                    d.writerow({"postcode": k, "longitude": v[0], "latitude": v[1]})

    def importXLS(self, file, sheet, column, header):
        # Read in data from xlsx, convert to coords, then write back to a csv
        wb = load_workbook(filename=file)
        ws = wb[sheet]

        postcodes = [cell.value for cell in ws[column]]

        if header:
            del postcodes[0]

        self.convertPostcodes(postcodes)

    def convertPostcodes(self, postcodes):
        for postcode in postcodes:
            coord = self.coordFromPostcode(postcode)
            if coord:
                self.coords.append(coord)

        self.saveCache()

    def coordFromPostcode(self, postcode):
        key = postcode.replace(" ", "")

        if key in self.cache:
            return self.cache[key]

        else:
            request = requests.get(self.endpoint + urllib.parse.quote_plus(postcode))
            response = request.json()

            if response["status"] == "match":
                coords = (
                    float(response["data"]["longitude"]),
                    float(response["data"]["latitude"]),
                )

                self.cache[key] = coords
                self.recache = True

                return coords
            else:
                return False

    def importPostcodes(file, sheet, column, header):
        pass

    def addMarkers(self, markers):
        self.markers.append(numpy.array(markers))

    def makeMap(self, **kwargs):
        sizes = {
            "a3": (11.693, 16.535),  # a3 portrait
            "a4": (8.268, 11.693),  # a4 portrait
        }

        paper = sizes[kwargs.get("paper", "a4")]
        orientation = kwargs.get("orientation", "landscape")
        oriented = (
            (paper[0], paper[1])
            if (orientation == "portrait")
            else (paper[1], paper[0])
        )

        self.figure = pl.figure(figsize=oriented)

        request = cimgt.QuadtreeTiles()
        ax = pl.axes(projection=request.crs)
        ax.add_image(request, self.tiledepth)

        if self.title:
            pl.title(label=self.title)

        if len(self.coords) > 0:
            x = numpy.asarray([m[0] for m in self.coords if m])
            y = numpy.asarray([m[1] for m in self.coords if m])

            xynps = ax.projection.transform_points(ccrs.Geodetic(), x, y)
            m = ax.hist2d(
                xynps[:, 0],
                xynps[:, 1],
                bins=[self.xbins, self.ybins],
                alpha=0.5,
                cmap=pl.cm.jet,
                cmin=1,
            )

            bounds = list(range(1, 10))
            pl.colorbar(
                m[3],
                ax=ax,
                shrink=0.4,
                format="%.0f",
                ticks=bounds,
                label="People per bin",
            )

        if len(self.markers) > 0:
            for markers in self.markers:
                unps = ax.projection.transform_points(
                    ccrs.Geodetic(),
                    markers[:, 1].astype(float),
                    markers[:, 0].astype(float),
                )
                pl.scatter(unps[:, 0], unps[:, 1], s=30, c=markers[:, 2], marker="D")

        pl.tight_layout(pad=4)

        file = kwargs.get("file", None)
        if file:
            self.saveMap(file)

    def saveMap(self, filename):
        if self.figure:
            self.figure.savefig(
                filename,
                dpi=self.figure.dpi,
            )
